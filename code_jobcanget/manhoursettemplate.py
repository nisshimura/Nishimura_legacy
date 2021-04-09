import csv 
import pathlib
import openpyxl
from datetime import datetime
import re 
import argparse

class Settemp(): 
    
    def args(self):
        parser = argparse.ArgumentParser(description='')

        parser.add_argument('input', help='inputcsv')
        parser.add_argument('output', help='outputexcel')
        parser.add_argument('start', help='y/m/d')
        parser.add_argument('end', help='y/m/d')

        args = parser.parse_args()
        return args

    def opencsv(self, manhourcsv):
            with open(manhourcsv,encoding='utf-8') as f:
                reader = csv.reader(f)
                jobcan = [row for row in reader]
        
            self.jobcan = jobcan

    def openexcel(self, resultexcel):
        wb = openpyxl.load_workbook(resultexcel)
        ws = wb['manhour']

        self.wb = wb
        self.ws = ws 

    def setdatas(self, start, end):
        jobcan = self.jobcan
        ws = self.ws

        count = 0

        for index, joblist in enumerate(jobcan):
            for i, j in enumerate(joblist):
                if i==0 and not index==0: 
                    correct = self.judge(start, end, j)
                    if correct == True:
                        c1 = ws.cell(row=index+1+count, column=i+1)
                        c1.value = j
                    else:
                        count -= 1
                        break
                else:
                    if i == 10:
                        c1 = ws.cell(row=index+1+count, column=i+1)
                        c1.value = str(j) + ':00'
                        c1.number_format = 'h:mm'
                    else:
                        c1 = ws.cell(row=index+1+count, column=i+1)
                        c1.value = j                        
    def deldatas(self):
        ws = self.ws

        maxrow = ws.max_row
        maxcolumn = ws.max_column
        minrow = ws.min_row
        mincolumn = ws.min_column
        
        for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincolumn, max_col=maxcolumn):
            for cell in row:
                cell.value = ''
            
    def savexlsx(self, outputxlsx):
        wb = self.wb
        wb.save(outputxlsx)
    
    def judge(self, start, end, obje):
        start = datetime.strptime(start, '%Y/%m/%d')
        end  = datetime.strptime(end, '%Y/%m/%d')
        obje = datetime.strptime(obje, '%Y/%m/%d')

        difs = obje - start
        dife = end - obje

        if int(difs.days) >= 0 and int(dife.days) >=0:
            return True
        else:
            return False

def main(editer):
    if editer==True:
        start = '2020/09/01'
        end = '2020/10/05'

        inputcsv = pathlib.Path("./jobcancsv/manhour_2020120216545fc7483510b52.csv")
        outputxlsx = pathlib.Path("./output/manhour_" + str(start.replace('/', '')) + '_' + str(end.replace('/', '')) + '.xlsx')
        copyxlsx = pathlib.Path("./output/manhour_template.xlsx")

        Job = Settemp()
        Job.opencsv(inputcsv)
        Job.openexcel(copyxlsx)
        Job.deldatas()
        Job.setdatas(start, end)
        Job.savexlsx(outputxlsx)
    else:
        copyxlsx = pathlib.Path("./output/manhour_template.xlsx")

        Job = Settemp()
        args = Job.args()
        Job.opencsv(args.input)
        Job.openexcel(copyxlsx)
        Job.deldatas()
        Job.setdatas(args.start, args.end)
        Job.savexlsx(args.output)
        print('SUCCESS')

if __name__ == "__main__":
    main(False) ####vscode上True、コマンドライン上False
#python C:\Users\github\workspace\jobcan_autotool\code_jobcanget\manhoursettemplate.py C:\Users\github\workspace\jobcan_autotool\code_jobcanget\jobcancsv\manhour_2020120217435fc753a5e0b12.csv C:\Users\github\workspace\jobcan_autotool\code_jobcanget\output\manhour_test.xlsx 2020/09/01 2020/10/05
