import csv
import openpyxl
import os 
import math
import re
import pathlib

class Jobs():

    anken = []  

    def opencsv(self, jobcancsv, boardcsv):
            with open(jobcancsv,encoding='utf-8') as f:
                reader = csv.reader(f)
                jobcan = [row for row in reader]

            with open(boardcsv, encoding='cp932') as g:
                reader = csv.reader(g)
                board = [row for row in reader]

            self.jobcan = jobcan
            self.board = board

    def openexcel(self, resultfile, resultexcel):
        if resultexcel.exists() == True:
            wb = openpyxl.load_workbook(resultexcel)
        else:
            resultfile.mkdir()
            wb = openpyxl.Workbook()

        ws = wb.worksheets[0]        
        ws.title = 'Jobcan_board_teamstats'
        
        self.wb = wb
        self.ws = ws
    
    def outputexcel(self):
        ws = self.ws
        board = self.board
        
        for rowindex, i in enumerate(board, start=1):
            for columnindex, j in enumerate(i, start=1):
                if columnindex == 1 and not rowindex == 1 :
                    self.anken.append(j)
                c1 = ws.cell(row=rowindex,column=columnindex)
                c1.value = j


        self.columnnum = len(board[0])  

    def inspect(self):
        jobcan = self.jobcan
        person = [] ##[['姓 名', 'プロジェクトコード', '工数'], ..., ['木村 康之', '1', '1:00']]
        person_dic = {}
        personnames = []

        for i in jobcan:
            no = i[6].lstrip('0')
            person.append([i[2], no, i[10]])
            person_dic[i[1]] = i[2]

        del person_dic['スタッフコード']
        person_dic = dict(sorted(person_dic.items()))
        
        for j in person_dic.values():
            personnames.append(j)

        person.remove(person[0])

        self.person = person
        self.names = personnames
    
    def add_nametag(self):
        names = self.names
        ws = self.ws
        culumnnum = self.columnnum
        self.name_start = int(culumnnum) + 2

        for index, i in enumerate(names, start=self.name_start):
            c1 = ws.cell(row=1, column=index)
            c1.value = i

    def make_calc_personsum(self):
        person = self.person     #person = [[name,ankenno,time],[name,ankenno,time]]
        
        dict_person = {}

        for i in person:
            if tuple([i[0],i[1]]) in dict_person:
                manhour = []
                manhour.append(dict_person[tuple([i[0],i[1]])])
                manhour.append(i[2])
                sumtime = self.calc(manhour)
                dict_person[tuple([i[0],i[1]])] = sumtime
            else:
                dict_person[tuple([i[0],i[1]])] = i[2]
        
        self.dict_person = dict_person

    def add_sumtimes(self):

        self.make_calc_personsum()
        
        ws = self.ws
        anken = self.anken                 #anken = [no,no,no]
        person = self.person
        names = self.names
        person_sum = self.dict_person

        for j in person:
            try:
                c2 = ws.cell(row=int(anken.index(j[1])) + 2, column=int(names.index(j[0])) + self.name_start)
                c2.value = person_sum[tuple([j[0],j[1]])]
            except:
                pass
        
        return True

    def add_allsum(self):
        names = self.names
        ws = self.ws
        anken = self.anken  

        for i in range(len(anken)):
            allsum_list = []
            c2 = ws.cell(row=i + 1, column=len(names) + self.name_start)
            if i == 0:
                c2.value = '合計'
            else:
                for j in range(len(names)):
                    c3 = ws.cell(row=i+1, column=j + self.name_start)
                    allsum_list.append(c3.value)
                c2.value = self.calc(allsum_list)
          
    def save_excel(self, resultexcel):
        wb = self.wb
        wb.save(resultexcel)

    def calc(self, time_list):
        hour = 0
        minute = 0
        for i in time_list:
            try:
                k = re.split(r':',i)
                hour += int(k[0])
                minute += int(k[1])
            except:
                pass
            
        if minute >= 60:
            hour += math.floor(minute/60)
            minute = minute % 60
        
        time = str(hour) + ':' + str(minute)
        return time

def main():
    jobcancsv = pathlib.Path('csv_datas/manhour.csv')
    boardcsv = pathlib.Path('csv_datas/board.csv')
    resultfile = pathlib.Path('excel_datas')
    resultexcel = resultfile / 'sophia_sample.xlsx'
    
    work = Jobs()
    work.opencsv(jobcancsv, boardcsv)
    work.openexcel(resultfile, resultexcel)
    work.outputexcel()
    work.inspect()
    work.add_nametag()
    work.add_sumtimes()
    work.add_allsum()
    work.save_excel(resultexcel)


if __name__ == "__main__":
    main()